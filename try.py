import openpyxl as opxl
import tkinter as tk
import os

wb = opxl.load_workbook("Expense.xlsx")
main = tk.Tk()

sheet = wb.active

#function
def totl():
    total = 0
    for x in range(1,sheet.max_row+1):
        cel = sheet.cell(row = x,column = 2)
        total = total + cel.value
        a = " : " + str(total)
    t = tk.Label(main,text = a).grid(row=3,column = 1)
def prn():
    b = e1.get()
    a = int(e2.get())
    r=sheet.max_row+1
    col = sheet.cell(row = r,column = 1)
    col.value = b
    col = sheet.cell(row = r,column = 2)
    col.value = a
    wb.save(filename =  "Expense.xlsx")
    print("File saved")

#label
tk.Label(main,text = "Item Name : ").grid(row = 0,column = 0)
tk.Label(main,text = "price : ").grid(row = 1,column = 0)

#entryS
e1 = tk.Entry(main)
e2 = tk.Entry(main)
e1.grid(row = 0,column = 1)
e2.grid(row = 1,column = 1)

#button
tk.Button(main,text = "Total",command = totl).grid(row = 3 , column = 0)
tk.Button(main,text = "Add to file",command = prn).grid(row = 2 , column = 0)
input()
